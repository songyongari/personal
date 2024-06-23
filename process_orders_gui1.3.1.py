import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

def process_orders(input_file, output_file):
    try:
        df = pd.read_excel(input_file, dtype={'주문번호': str})

        # 필요한 컬럼만 추출
        df_filtered = df[['주문번호', '상품명', '옵션정보', '클레임상태', '수량']]

        # 가격 정보 사전 정의 (한 줄로 요약)
        price_dict = {
            '도시락 14개 자유선택  (14개x1회 배송): 시그니처 1. 강남역 호랑이 삼겹' : 9500,
            '도시락 14개 자유선택  (14개x1회 배송): 시그니처 2. 수원 왕갈비통 닭목살' : 9500,
            '도시락 14개 자유선택  (14개x1회 배송): 시그니처 3. 기사식당 최강 제육' : 9500,
            '도시락 14개 자유선택  (14개x1회 배송): 시그니처 4. 춘천 들깨 닭갈비' : 9500,
            '도시락 14개 자유선택  (14개x1회 배송): 시그니처 5. 수랏간 삼치 솥밥' : 9500,
            '도시락 14개 자유선택  (14개x1회 배송): 시그니처 6. 항아리 차돌 된장' : 9500,
            '도시락 14개 자유선택  (14개x1회 배송): 오리지널 1. 수비드 통삼겹 된장덮밥' : 8500,
            '도시락 14개 자유선택  (14개x1회 배송): 오리지널 2. 수비드 통삼겹 들기름 막국수' : 8500,
            '도시락 14개 자유선택  (14개x1회 배송): 오리지널 3. 훈제오리 들깨 크림 리조또' : 8500,
            '도시락 14개 자유선택  (14개x1회 배송): 오리지널 4. 우삼겹 규동' : 8500,
            '도시락 14개 자유선택  (14개x1회 배송): 오리지널 5. 우삼겹 오일 파스타' : 8500,
            '도시락 14개 자유선택  (14개x1회 배송): 오리지널 6. B.T.S 치킨치즈 리조또' : 8500,
            '도시락 14개 자유선택  (14개x1회 배송): 오리지널 7. 소고기 버섯 들깨 덮밥' : 8500,
            '도시락 14개 자유선택  (14개x1회 배송): 오리지널 8. 저당 두부면 라자냐' : 8500,
            '도시락 14개 자유선택  (14개x2회 배송): 시그니처 1. 강남역 호랑이 삼겹' : 19000,
            '도시락 14개 자유선택  (14개x2회 배송): 시그니처 2. 수원 왕갈비통 닭목살' : 19000,
            '도시락 14개 자유선택  (14개x2회 배송): 시그니처 3. 기사식당 최강 제육' : 19000,
            '도시락 14개 자유선택  (14개x2회 배송): 시그니처 4. 춘천 들깨 닭갈비' : 19000,
            '도시락 14개 자유선택  (14개x2회 배송): 시그니처 5. 수랏간 삼치 솥밥' : 19000,
            '도시락 14개 자유선택  (14개x2회 배송): 시그니처 6. 항아리 차돌 된장' : 19000,
            '도시락 14개 자유선택  (14개x2회 배송): 오리지널 1. 수비드 통삼겹 된장덮밥' : 17000,
            '도시락 14개 자유선택  (14개x2회 배송): 오리지널 2. 수비드 통삼겹 들기름 막국수' : 17000,
            '도시락 14개 자유선택  (14개x2회 배송): 오리지널 3. 훈제오리 들깨 크림 리조또' : 17000,
            '도시락 14개 자유선택  (14개x2회 배송): 오리지널 4. 우삼겹 규동' : 17000,
            '도시락 14개 자유선택  (14개x2회 배송): 오리지널 5. 우삼겹 오일 파스타' : 17000,
            '도시락 14개 자유선택  (14개x2회 배송): 오리지널 6. B.T.S 치킨치즈 리조또' : 17000,
            '도시락 14개 자유선택  (14개x2회 배송): 오리지널 7. 소고기 버섯 들깨 덮밥' : 17000,
            '도시락 14개 자유선택  (14개x2회 배송): 오리지널 8. 저당 두부면 라자냐' : 17000,
            '시그니처: 1. 강남역 호랑이 삼겹' : 9500,
            '시그니처: 2. 수원 왕갈비통 닭목살' : 9500,
            '시그니처: 3. 기사식당 최강 제육' : 9500,
            '시그니처: 4. 춘천 들깨 닭갈비' : 9500,
            '시그니처: 5. 수랏간 삼치 솥밥' : 9500,
            '시그니처: 6. 항아리 차돌 된장' : 9500,
            '시그니처: 도시락 6종 x  1팩' : 57000,
            '시그니처: 도시락 6종 x 1팩 (조합 선택 불가)' : 57000,
            '오리지널: 1. 수비드 통삼겹 된장덮밥' : 8500,
            '오리지널: 2. 수비드 통삼겹 들기름 막국수' : 8500,
            '오리지널: 3. 훈제오리 들깨 크림 리조또' : 8500,
            '오리지널: 4. 우삼겹 규동' : 8500,
            '오리지널: 5. 우삼겹 오일 파스타' : 8500,
            '오리지널: 6. B.T.S 치킨치즈 리조또' : 8500,
            '오리지널: 7. 소고기 버섯 들깨 덮밥' : 8500,
            '오리지널: 8. 저당 두부면 라자냐' : 8500,
            '오리지널: 도시락 8종 x 1팩 (조합 선택 불가)' : 68000,
            '패키지 선택: [구독 패키지 1] 황금비율 1:2:7' : 238000,
            '패키지 선택: [구독 패키지 2] 실패없는 베스트셀러' : 246000,
            '패키지 선택: [구독 패키지 3] 김윤겸이 매일 먹는 도시락' : 244000
        }

        # 취소 및 반품 항목 제외
        df_filtered = df_filtered[~df_filtered['클레임상태'].isin(['취소완료', '반품완료'])]

        # 첫 번째 시트: 매출 요약 테이블
        # 일반 매출 (챌린지 신청 및 신청 안함 항목을 제외한 매출)
        일반_매출 = df_filtered[
            ~df_filtered['상품명'].str.contains('챌린지 신청|신청 안함', na=False)
        ]
        일반_매출['단가'] = 일반_매출['옵션정보'].map(price_dict)
        일반_매출['금액'] = 일반_매출['단가'] * 일반_매출['수량']
        일반_매출['금액'] = 일반_매출['금액'].fillna(0)
        일반_매출_total = 일반_매출.groupby('주문번호').agg({'상품명': 'first', '금액': 'sum', '클레임상태': 'first'}).reset_index()
        일반_매출_합계 = 일반_매출_total['금액'].sum()

        def create_summary_sheet(writer, 일반_매출_합계, 챌린지_미신청_매출, 챌린지_신청_매출):
            챌린지_총_매출 = 챌린지_미신청_매출 + 챌린지_신청_매출
            전체_도시락_매출_합계 = 일반_매출_합계 + 챌린지_총_매출

            summary_data = {
                '일반 도시락 매출': [일반_매출_합계, 0, 0, 일반_매출_합계],
                '챌린지 도시락 매출': [챌린지_총_매출, 챌린지_미신청_매출, 챌린지_신청_매출, 챌린지_총_매출],
                '전체 도시락 매출': [전체_도시락_매출_합계, 0, 0, 전체_도시락_매출_합계]
            }

            summary_df = pd.DataFrame(summary_data, index=['매출', '미신청 매출', '신청 매출', '비율'])
            summary_df.loc['비율'] = summary_df.loc['매출'] / 전체_도시락_매출_합계 * 100
            summary_df.loc['비율'] = summary_df.loc['비율'].map('{:.0f}%'.format)

            summary_df.iloc[0:3] = summary_df.iloc[0:3].applymap(lambda x: f"\\{x:,.0f}" if isinstance(x, (int, float)) else x)

            summary_df.to_excel(writer, sheet_name='매출 요약')


            # 두 번째 테이블: 챌린지 매출과 부트캠프 매출
            챌린지_주문수 = df_filtered[df_filtered['상품명'].str.contains('챌린지 신청', na=False)]['주문번호'].nunique()
            부트캠프_주문수 = df_filtered[df_filtered['옵션정보'].str.contains('부트캠프 신청', na=False)]['주문번호'].nunique()

            두번째_테이블_data = {
                '챌린지 매출': [50000, 챌린지_주문수, 챌린지_주문수 * 50000],
                '부트캠프 매출': [399000, 부트캠프_주문수, 부트캠프_주문수 * 399000]
            }

            두번째_테이블_df = pd.DataFrame(두번째_테이블_data, index=['금액', '수량', '매출'])

            두번째_테이블_df.iloc[0::2] = 두번째_테이블_df.iloc[0::2].applymap(lambda x: f"\\{x:,.0f}" if isinstance(x, (int, float)) else x)

            두번째_테이블_df.to_excel(writer, sheet_name='매출 요약', startrow=8)

        # 두 번째 시트: 챌린지 매출 (챌린지 신청, 신청 안함)
        def create_challenge_sales_sheet(writer, df_filtered, price_dict):
            # '챌린지'가 포함된 모든 주문 가져오기
            challenge_related_orders = df_filtered[df_filtered['상품명'].str.contains('챌린지', na=False)].copy()

            # '챌린지 신청'과 '신청 안함' 분리
            challenge_orders = challenge_related_orders[challenge_related_orders['상품명'].str.contains('챌린지 신청', na=False)]
            non_challenge_orders = challenge_related_orders[~challenge_related_orders['상품명'].str.contains('챌린지 신청', na=False)]

            # 해당 주문번호로 모든 주문 정보 조회
            challenge_order_numbers = challenge_orders['주문번호'].unique()
            non_challenge_order_numbers = non_challenge_orders['주문번호'].unique()

            all_challenge_orders = df_filtered[df_filtered['주문번호'].isin(challenge_order_numbers)].copy()
            all_non_challenge_orders = df_filtered[df_filtered['주문번호'].isin(non_challenge_order_numbers)].copy()

            all_challenge_orders['단가'] = all_challenge_orders['옵션정보'].map(price_dict)
            all_non_challenge_orders['단가'] = all_non_challenge_orders['옵션정보'].map(price_dict)

            all_challenge_orders['금액'] = all_challenge_orders['단가'] * all_challenge_orders['수량']
            all_non_challenge_orders['금액'] = all_non_challenge_orders['단가'] * all_non_challenge_orders['수량']

            all_challenge_orders['금액'] = all_challenge_orders['금액'].fillna(0)
            all_non_challenge_orders['금액'] = all_non_challenge_orders['금액'].fillna(0)

            # '챌린지 신청'에 해당하는 주문은 '상품명'에서 가져오기
            all_challenge_orders.loc[all_challenge_orders['주문번호'].isin(challenge_order_numbers), '상품명'] = \
                all_challenge_orders.loc[all_challenge_orders['상품명'].str.contains('챌린지 신청', na=False), '상품명']
            all_non_challenge_orders['상품명'] = '신청 안함'

            challenge_total = all_challenge_orders.groupby('주문번호').agg({'상품명': 'first', '금액': 'sum', '클레임상태': 'first'}).reset_index()
            non_challenge_total = all_non_challenge_orders.groupby('주문번호').agg({'상품명': 'first', '금액': 'sum', '클레임상태': 'first'}).reset_index()

            result = pd.concat([challenge_total, non_challenge_total], ignore_index=True)
            result.to_excel(writer, sheet_name='챌린지 매출', index=False)

            return challenge_total, non_challenge_total


        # 세 번째 시트: 부트캠프 매출
        def create_bootcamp_sales_sheet(writer, df_filtered):
            bootcamp_orders = df_filtered[df_filtered['옵션정보'].str.contains('부트캠프 신청', na=False)].copy()
            bootcamp_orders = bootcamp_orders[['주문번호', '상품명', '옵션정보', '클레임상태']]
            bootcamp_orders.to_excel(writer, sheet_name='부트캠프 매출', index=False)

            return bootcamp_orders

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 챌린지 매출 계산
            challenge_total, non_challenge_total = create_challenge_sales_sheet(writer, df_filtered, price_dict)
            챌린지_유입_매출 = non_challenge_total['금액'].sum()
            챌린지_신청_매출 = challenge_total['금액'].sum()

            # 매출 요약 시트 생성
            create_summary_sheet(writer, 일반_매출_합계, 챌린지_유입_매출, 챌린지_신청_매출)
            
            # 부트캠프 매출 시트 생성
            create_bootcamp_sales_sheet(writer, df_filtered)

        # 시트 순서 재조정
        wb = load_workbook(output_file)
        sheet_order = ['매출 요약', '챌린지 매출', '부트캠프 매출']
        wb._sheets = [wb[sheet] for sheet in sheet_order if sheet in wb.sheetnames]
        wb.save(output_file)

        messagebox.showinfo("성공", f"파일이 성공적으로 생성되었습니다: {output_file}")

    except Exception as e:
        messagebox.showerror("오류", f"오류가 발생했습니다: {e}")

def select_input_file():
    input_file = filedialog.askopenfilename(title="입력 파일 선택", filetypes=[("Excel files", "*.xlsx *.xls")])
    input_entry.delete(0, tk.END)
    input_entry.insert(0, input_file)

def select_output_file():
    output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", title="출력 파일 저장", filetypes=[("Excel files", "*.xlsx")])
    output_entry.delete(0, tk.END)
    output_entry.insert(0, output_file)

def run_process():
    input_file = input_entry.get()
    output_file = output_entry.get()
    if not input_file or not output_file:
        messagebox.showwarning("경고", "입력 파일과 출력 파일을 모두 선택해야 합니다.")
        return
    process_orders(input_file, output_file)

# GUI 설정
root = tk.Tk()
root.title("Process Orders")

tk.Label(root, text="입력 파일:").grid(row=0, column=0, padx=10, pady=10)
input_entry = tk.Entry(root, width=50)
input_entry.grid(row=0, column=1, padx=10, pady=10)
tk.Button(root, text="파일 선택", command=select_input_file).grid(row=0, column=2, padx=10, pady=10)

tk.Label(root, text="출력 파일:").grid(row=1, column=0, padx=10, pady=10)
output_entry = tk.Entry(root, width=50)
output_entry.grid(row=1, column=1, padx=10, pady=10)
tk.Button(root, text="파일 저장", command=select_output_file).grid(row=1, column=2, padx=10, pady=10)

tk.Button(root, text="실행", command=run_process).grid(row=2, column=0, columnspan=3, pady=20)

root.mainloop()
